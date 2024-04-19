import openpyxl
from bs4 import BeautifulSoup
import csv
from tqdm import tqdm


def _extract_unique_values_from_tsv(DAT_tsv_file: str,  # noqa
                                    DAT_header_htm_file: str,  # noqa
                                    data_dictionary_xls: dict,
                                    data_dictionary_sheet_name_for_dat_file: str,
                                    max_unique: int = 25,
                                    name_of_sheet_with_dictionary: str = 'Flatfile Formats'
                                    ) -> dict:
    """
    Extract unique values from a TSV file based on var_name defined in an HTM file, and enhance the unique values
    with formatted values and data types from the data dictionary.

    :param DAT_tsv_file: Path to the TSV data file.
    :param DAT_header_htm_file: Path to the HTM file containing data row definitions.
    :param data_dictionary_xls: Parsed data with format information and mappings.
    :param data_dictionary_sheet_name_for_dat_file: Name of the sheet corresponding to the DAT file
    :param max_unique: Maximum number of unique values to store per column.
    :param name_of_sheet_with_dictionary: Name of the sheet that contains the data dictionary.
    :return: Dictionary with column names as keys and lists of unique values as values.
    """
    # open xls file
    wb = openpyxl.load_workbook(data_dictionary_xls)

    # First, get all the variables from the sheet of interest
    data_dictionary_sheet_for_dat_file = wb[data_dictionary_sheet_name_for_dat_file]

    data_dictionary_for_dat = {}
    for row in tqdm(data_dictionary_sheet_for_dat_file.iter_rows(min_row=3,
                                                                 values_only=True),
                    desc=f"Parsing data dictionary in {data_dictionary_sheet_name_for_dat_file} sheet", unit="rows"):
        var_name, description, form, var_start_date, var_end_date, form_section, data_type, sas_analysis_format, comment, *_ = row
        if var_name in data_dictionary_for_dat:
            raise ValueError(f"Duplicate variable name: {var_name}")
        data_dictionary_for_dat[var_name] = {
            'description': description,
            'form': form,
            'var_start_date': str(var_start_date),
            'var_end_date': str(var_end_date),
            'form_section': form_section,
            'data_type': data_type,
            'sas_analysis_format': sas_analysis_format,
            'comment': comment
        }

    #
    # Next get all observed values in the DAT file
    #
    # Parse HTM file to extract var_name using the first <td> of each <tr> in <tbody>
    with open(DAT_header_htm_file, 'r', encoding='windows-1252') as file:
        soup = BeautifulSoup(file, 'html.parser')
        var_name = [tr.find('td').text.strip() for tr in soup.select('tbody tr') if tr.find('td')]

    # Initialize a dictionary to collect unique observed values
    observed_values = {this_var: set() for this_var in var_name}

    # Read TSV file and extract unique values with a limit
    with open(DAT_tsv_file, 'r', encoding='utf-8') as file:
        reader = csv.DictReader(file, fieldnames=var_name, delimiter='\t')
        for row in tqdm(reader, desc='Extracting unique observed values', unit='rows'):
            for header in var_name:
                if len(observed_values[header]) < max_unique:
                    observed_values[header].add(row[header])

    # now put observed values into the data dictionary
    for var_name, var_data in tqdm(data_dictionary_for_dat.items(),
                                   desc='Adding observed values to data dictionary', unit='variables'):
        if var_name in observed_values:
            if 'observed_values' in var_data:
                raise ValueError(f"Duplicate observed values for {var_name}")
            else:  # Add observed values to the data dictionary
                var_data['observed_values'] = list(observed_values[var_name])

    # Read and parse the XLS data dictionary sheet ("Flatfile Formats")
    # This sheet contains the data dictionary with valid values and data types.
    # Weirdly, the variables in the data dictionary are slightly different from
    # those in the DAT file
    sheet = wb[name_of_sheet_with_dictionary]
    SAS_format_data_dictionary = {}
    for row in tqdm(sheet.iter_rows(min_row=3, values_only=True),
                    desc='Parsing data dictionary', unit='rows'):
        var_name, data_field_value, data_field_formatted_value, data_type, *_ = row
        if var_name not in SAS_format_data_dictionary:
            SAS_format_data_dictionary[var_name] = {
                'valid_values': {},
                'data_type': data_type
            }
        SAS_format_data_dictionary[var_name]['valid_values'][
            data_field_value] = data_field_formatted_value

    # loop through data dictionary and add SAS format data to the data dictionary
    for var_name, var_data in data_dictionary_for_dat.items():
        sas_format_name = var_data['sas_analysis_format']
        if sas_format_name in SAS_format_data_dictionary:
            var_data['valid_values'] = SAS_format_data_dictionary[sas_format_name]['valid_values']
            var_data['data_type'] = SAS_format_data_dictionary[sas_format_name]['data_type']

    # Return the data dictionary
    return data_dictionary_for_dat

